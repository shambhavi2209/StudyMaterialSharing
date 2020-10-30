
"""
1. Install all these packages below first.

2. Second thing you have to do is search here for C:\\Users\\hunny\\Documents and replace the location with where you have paste the folder, thats it.
   then run the code freely ,ENJOY.

3. Read all the comments with code which will make you understand that what happens with that specific script.   """


from tkinter import * 
from tkinter import ttk
from PIL import Image,ImageTk
from tkinter import messagebox 
import pandas as pd
import xlrd
from openpyxl import *
import pyttsx3
import PyPDF2
import keyboard
import os

class study:

    def __init__(self,root):

        self.root = root
        self.root.wm_iconbitmap("images/icon.ico") #Sets favicon to the screen
        self.root.title("Study Material Sharing Application")

        #--Window
        self.root.geometry("1100x700+400+100")
        self.root.minsize(1100,700)
        self.root.maxsize(1100,700)
        self.root.resizable(False,False)

        #--Background Image
        self.bg = ImageTk.PhotoImage(file = "images/two.jpg")
        self.bg_fit = Label(self.root,image = self.bg).place(x=0,y=0,relwidth=1,relheight=1)
        title = Label(self.root, text="Study Material application", font="comicsansms 35 bold", fg="white",bg = "#2f2641",relief = SUNKEN,borderwidth = 10)
        title.pack(pady = 90,ipadx = 5,ipady = 5)

        #Account Frame
        account_frame = Frame(self.root,bg = "white",relief = RAISED,borderwidth = 30)
        account_frame.place(x = 190,y = 200,height = 440,width = 500)

        #Login page
        login_title = Label(account_frame,text = "Login",font = "Impact 35 bold",fg = "#2f2641",bg = "white").place(x=90,y=30)
        span= Label(account_frame, text="Login to access", font="comicsansms 10 bold", fg="#d25d17", bg="white").place(x=95,y=100)
        self.name = StringVar()
        self.password = StringVar()

        username = Label(account_frame, text="Username..", font="comicsansms 15 bold", fg="#4d4d4d", bg="white").place(x=40, y=140)
        self.user_entry = Entry(account_frame,textvariable = self.name,font = ("times new roman",15),bg = "lightgray",relief = SUNKEN,borderwidth = 6)
        self.user_entry.place(x=45,y=173,width = 350,height = 35)

        password = Label(account_frame, text="Password..", font="comicsansms 15 bold", fg="#4d4d4d", bg="white").place(x=40, y=215)
        self.user_pass = Entry(account_frame, textvariable = self.password,show="*",font=("times new roman", 15), bg="lightgray",relief = SUNKEN,borderwidth = 6)
        self.user_pass.place(x=45, y=247, width=350, height=35)

        new_user = Button(account_frame,text = "New User?",cursor ="hand2",command = self.signup,bg = "white",fg = "#d25d17",font=("times new roman", 12),relief = FLAT).place(x = 230,y = 287)
        Log_button = Button(account_frame,command = self.login,cursor ="hand2",text="Login", fg="white", bg="#d25d17", font=("times new roman", 20),relief=GROOVE).place(x=80, y=290,height=40)

    def signup(self):
        self.root = root
        self.top = Toplevel()  # Same as Tk() but it is used for creating window over window with all functions same.
        self.top.wm_iconbitmap("images/icon.ico")  # Favicon
        self.top.title("Signup")
        self.top.geometry("500x450+590+340")
        self.top.minsize(500, 450)
        self.top.maxsize(500, 450)
        self.top.resizable(False, False)

        signup_frame = Frame(self.top, bg="white", relief=RAISED, borderwidth=30)
        signup_frame.place(x=0, y=0, height=450, width=500)

        signup_title = Label(signup_frame, text="Signup", font="Impact 35 bold", fg="#2f2641", bg="white").place(x=90,
                                                                                                                 y=30)
        span = Label(signup_frame, text="Make your account", font="comicsansms 10 bold", fg="#d25d17",
                     bg="white").place(x=95, y=100)

        self.sign_name = StringVar()
        self.sign_password = StringVar()

        username = Label(signup_frame, text="Set username", font="comicsansms 15 bold", fg="#4d4d4d", bg="white").place(
            x=40, y=140)
        self.user_entry = Entry(signup_frame, textvariable=self.sign_name, font=("times new roman", 15), bg="lightgray",
                                relief=SUNKEN, borderwidth=6)
        self.user_entry.place(x=45, y=173, width=350, height=35)

        password = Label(signup_frame, text="Make password", font="comicsansms 15 bold", fg="#4d4d4d",
                         bg="white").place(x=40, y=215)
        pass_span = Label(signup_frame, text="(Must contains characters combination)", font="comicsansms 8 bold",
                          fg="#d25d17", bg="white").place(x=45, y=247)
        self.user_pass = Entry(signup_frame, textvariable=self.sign_password, show="*", font=("times new roman", 15),
                               bg="lightgray", relief=SUNKEN, borderwidth=6)
        self.user_pass.place(x=45, y=270, width=350, height=35)

        Button(signup_frame, text="If exist", cursor="hand2", command=self.top.destroy, bg="white", fg="#d25d17",
               font=("times new roman", 12), relief=FLAT).place(x=230, y=320)
        sign_button = Button(signup_frame, command=self.sign, cursor="hand2", text="Signup", fg="white", bg="#d25d17",
                             font=("times new roman", 20), relief=GROOVE).place(x=80, y=320, height=40)

    def sign(self):

        self.file = "data/data2.xlsx"
        self.filex = load_workbook(self.file)
        self.sh = self.filex["Sheet1"]
        self.row = self.sh.max_row
        isSign = False
        self.x = 1
        if self.sign_name.get() == "" or self.sign_password.get() == "":
            messagebox.showerror("Error", "Please fill all the fields", parent=self.root)
        else:
            isSign = True
            while (self.x <= (self.row + 1)):
                if (self.x == (self.row + 1)):
                    self.sh[f"A{self.row + 1}"] = self.sign_name.get()
                    self.sh[f"B{self.row + 1}"] = self.sign_password.get()

                self.x += 1
        self.filex.save(self.file)
        # print(self.sh.cell(self.sh.max_row,1))
        if (isSign):
            messagebox.showinfo("Welcome", f"Welcome {self.sign_name.get()},Proceed with Login now ", parent=self.root)

    #This below function calls when click Log_button
    def login(self):
        self.data = pd.ExcelFile("data/data2.xlsx")
        dd = self.data.parse()  #Fetch all the data from excel

        for x in dd.values:
            if self.name.get() == "" or self.password.get() == "":
                messagebox.showerror("Error", "Please fill all the fields", parent=self.root)
                break
            elif self.name.get() in x:
                if x[1]==self.password.get():
                    x = messagebox.showinfo("Welcome", f"Welcome {self.user_entry.get()} to our application", parent=self.root)
                    print(x)
                    if x=="ok":
                        self.mainwork()
        else:
            messagebox.showerror("Error", "Invalid credentials", parent=self.root)
        # Here for else statement works whenever for loop fails and stop else will work.



    #after account cresentials , all work starts from here
    def mainwork(self):
        self.root.destroy()
        self.newroot = Tk()
        self.newroot.wm_iconbitmap("images/icon.ico")
        self.newroot.title("Study Material Sharing Application")

        # --Window
        self.newroot.geometry("1100x800+400+100")
        self.newroot.minsize(1100, 800)
        self.newroot.maxsize(1100, 800)
        self.newroot.resizable(False, False)

        self.bg = ImageTk.PhotoImage(file="images/five.jpg")
        self.bg_fit = Label(self.newroot, image=self.bg).place(x=0, y=0, relwidth=1, relheight=1)
        title = Label(self.newroot, text="Choose any subjects you want..", font="comicsansms 35 bold", fg="white",
                      bg="#2f2641", relief=SUNKEN, borderwidth=10)
        title.pack(pady=30, ipadx=5, ipady=5)

        #Creating first canvas for above 4 subjects.
        c1 = Canvas(self.newroot, width=300, height=300, relief=RAISED, borderwidth=20, bg="grey")
        c1.pack()
        image1 = PhotoImage(file="images/s1.png")
        image2 = PhotoImage(file="images/s2.png")
        image3 = PhotoImage(file="images/s3.png")
        image4 = PhotoImage(file="images/s4.png")

        """Here question arises why lambda function is used? This is so becoz If we want to pass an arguments in the function then
        Simple function will not work , If you use simple function then whenever you open the window , that function will evoke immediately
        and will results in different ,thats why lambda function is used here."""

        b = Button(c1, image=image1, compound=LEFT, relief=FLAT, bg="grey", command=lambda:self.common("Python"))
        b.pack(side=LEFT, padx=23, pady=23)

        b = Button(c1, image=image2, compound=LEFT, relief=FLAT, bg="grey", command=lambda:self.common("Maths"))
        b.pack(side=LEFT, padx=23, pady=23)

        b = Button(c1, image=image3, compound=LEFT, relief=FLAT, bg="grey", command=lambda:self.common("C++"))
        b.pack(side=LEFT, padx=23, pady=23)

        b = Button(c1, image=image4, compound=LEFT, relief=FLAT, bg="grey", command=lambda:self.common("DSA"))
        b.pack(side=LEFT, padx=23, pady=23)

        #Creating second canvas for below 4 subjects.
        c2 = Canvas(self.newroot, width=300, height=300, relief=RAISED, borderwidth=20, bg="grey")
        c2.pack(pady=10)

        image21 = PhotoImage(file="images/s5.png")
        image22 = PhotoImage(file="images/s7.png")
        image23 = PhotoImage(file="images/s8.png")
        image24 = PhotoImage(file="images/s9.png")

        b2 = Button(c2, image=image21, compound=LEFT, relief=FLAT, bg="grey", command=lambda:self.common("DBMS"))
        b2.pack(side=LEFT, padx=23, pady=23)

        b2 = Button(c2, image=image22, compound=LEFT, relief=FLAT, bg="grey", command=lambda:self.common("English"))
        b2.pack(side=LEFT, padx=23, pady=23)

        b2 = Button(c2, image=image23, compound=LEFT, relief=FLAT, bg="grey", command=lambda:self.common("Java"))
        b2.pack(side=LEFT, padx=23, pady=23)

        b2 = Button(c2, image=image24, compound=LEFT, relief=FLAT, bg="grey", command=lambda:self.common("Software_engg"))
        b2.pack(side=LEFT, padx=23, pady=23)
        mainloop()

    def common(self,x):
        self.win = Toplevel()
        self.win.wm_iconbitmap("images/icon.ico")
        self.win.title(x)
        self.win.geometry("1000x600+450+300")
        self.win.minsize(1000, 600)
        self.win.maxsize(1000, 600)
        self.win.resizable(False, False)

        # From line 213 to 229, a function of scrollbar and its bind with window has written.
        main_frame = Frame(self.win)
        main_frame.pack(fill=BOTH, expand=1)

        my_canvas = Canvas(main_frame)
        my_canvas.pack(side=LEFT, fill=BOTH, expand=1)

        my_scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
        my_scrollbar.pack(side=RIGHT, fill=Y)

        my_canvas.configure(yscrollcommand=my_scrollbar.set)
        my_canvas.bind('<Configure>', lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all")))

        second_frame = Frame(my_canvas)

        my_canvas.create_window((0, 0), window=second_frame, anchor="nw")
        self.back = ImageTk.PhotoImage(file="images/back3.jpg")
        self.back_fit = Label(second_frame, image=self.back).place(x=0, y=0, relwidth=1, relheight=1)


        #After this above execution second_frame will act as a new frame as root for further execution

        title = Label(second_frame,text=x, font="Impact 35 bold", fg="#2f2641", bg="white")
        title.pack(pady=30, ipadx=5, ipady=5)

        #Lists below are created just to give a different variable names during run time between the loop such such every widget will get its identity.
        list1 = ["c1", "c2", "c3", "c4", "c5", "c6"]
        self.image1 = ["image1", "image2", "image3", "image4", "image5", "image6"]

        files = []
        btn = []

        files2 = []
        btn2 = []

        books_name = load_workbook(f"data/booksname/{x}.xlsx") #This excel file contains the titles of that specific book.
        show = books_name["Sheet1"]

        for i in range(1,7):
            files.append("View" + str(i))
        print(len(files))

        for i in range(1,7):
            files2.append("Intro" + str(i))
        # print(len(files2))

        for i in range(len(files)):
            book = show.cell(i+1, 1)
            list1[i] = Canvas(second_frame, width=320, height=300, relief=RAISED, borderwidth=20, bg="#add5ff")
            list1[i].pack(ipadx=100, padx=120)

            self.image1[i] = ImageTk.PhotoImage(file=f"images/books/{x}/{i+1}.png")
            self.down = ImageTk.PhotoImage(file="images/down.png")
            self.view = ImageTk.PhotoImage(file="images/view.png")

            Button(list1[i], image=self.image1[i]).pack(side=LEFT, pady=23, padx=15)

            Label(list1[i], text=f"{book.value}", font="comicsansms 18 bold", fg="black",bg="#6e81fa", relief=SUNKEN, borderwidth=10).pack(side=TOP, pady=60)

            btn.append(Button(list1[i], text=files[i], relief=RAISED, borderwidth=6, command=lambda c=i: os.startfile(f"data\\books\\{x}\\{btn[c].cget('text')}.pdf")))
            btn[i].pack(side=BOTTOM, padx=10, pady=40)

            btn2.append(Button(list1[i], text=files2[i], relief=RAISED, borderwidth=6, command=lambda c=i:self.speak_pdf(x,btn[c].cget('text'))))
            btn2[i].pack(side=BOTTOM, padx=23)

    def speak_pdf(self,x,y):
        # print(y)
        a = PyPDF2.PdfFileReader(f"data\\booksintro\\{x}\\{y}.pdf")
        str = "Hello Reader! The Small intro of this pdf is ! " + a.getPage(0).extractText() + "!!The More Details you will be find in this book, Do check it out."
        self.speak(str)

    def speak(self,audio):
        engine = pyttsx3.init('sapi5')
        voices = engine.getProperty('voices')
        engine.setProperty('voice', voices[0].id)
        engine.say(audio)
        engine.runAndWait()
        if keyboard.is_pressed('q'):
            engine.stop()


if __name__ == '__main__':

    """Execution of code starts from here by making an object of class study and
       passing root as an argument which calls the parameterized constructor of that class"""

    root = Tk()
    obj1 = study(root)
    mainloop()